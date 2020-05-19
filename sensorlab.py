from openpyxl import Workbook, load_workbook
from algorithm import Algorithm

PATH_DS = './downloads/example_dataset.xlsx'
SHEET = ''


def from_excel(path, sheet_name, rangex, rangey):
    wb = load_workbook(filename=path)
    sheet = wb[sheet_name]
    dataset = []
    for x, y in zip(sheet[rangex], sheet[rangey]):
        dataset.append((x[0].value, y[0].value))
    return dataset

def to_excel(path=None):
    # TODO: Write save dataset to .xlsx
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 42
    ws.append([1,2,3])
    wb.save('./downloads/sample.xlsx')


def plot(dataset_base, dataset_finish, frame='decart'):
    """ This function used to draw  """
    x = [x for (x, _) in dataset_base]
    y = [y for (_, y) in dataset_base]

    xc = [x for (x, _) in dataset_finish]
    yc = [y for (_, y) in dataset_finish]

    import matplotlib
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots()
    fig.set_size_inches(8.5, 8.5)
    ax.plot(x, y, marker='.')
    ax.plot(xc, yc, marker='.', color='red')
    ax.set(xlabel='B, uT', ylabel='C, uT',
           title='Magnetic ellips')
    ax.grid()
    fig.savefig("test.png")
    plt.show()


class Calibration:
    algorithm = {'maxdub': Algorithm}

    def __init__(self, fields, algorithm=None):
        if fields is not None:
            self._fields = fields

    def calculate(self):
        pass

    def correction(self, x, y):
        # <1> Formula calculation field X
        # <2> Formula calculation field Y
        return x, y


if __name__ == '__main__':
    dataset = from_excel(
        path='./downloads/example_dataset.xlsx',
        sheet_name='Лист6',
        rangex='H7:H51',
        rangey='I7:I51'
    )
    maxdub = Algorithm(dataset)
    ds_correct = [maxdub.correct(x,y) for (x,y) in dataset]
    print(f"\n{len(dataset)}, {dataset=}\n"
          f"{len(ds_correct)}, {ds_correct=}")

    print(maxdub)

    plot(dataset,ds_correct)
