import argparse
import csv

from openpyxl import load_workbook, Workbook

PATH_DS = '../downloads/example_dataset.xlsx'
SHEET = ''

# Excel
# TODO: Write convertor test data to csv file
def from_excel(path, sheet_name, rangex, rangey):
	wb = load_workbook(filename=path)
	sheet = wb[sheet_name]
	dataset = []
	for x, y in zip(sheet[rangex], sheet[rangey]):
		dataset.append((x[0].value, y[0].value))
	return dataset


# TODO: Write save dataset to .xlsx
def to_excel(path=None):
	wb = Workbook()
	ws = wb.active
	ws['A1'] = 42
	ws.append([1, 2, 3])
	wb.save('./downloads/sample.xlsx')


# CSV
def from_csv(fname):
	with open(fname, "r") as f:
		reader = csv.reader(f, quoting=csv.QUOTE_NONNUMERIC)
		return [row for row in reader]


def to_csv(fname, data, mode='w'):
	with open(fname, mode) as f:
		writer = csv.writer(f, delimiter=',')
		writer.writerows(data)


def get_arguments():
	p = argparse.ArgumentParser()
	p.add_argument('--mode', action='store', dest='mode', default=None)
	p.add_argument('--data', action='store', dest='path_to_dataset', default=None)
	return p.parse_args()
